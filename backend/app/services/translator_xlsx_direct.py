import openpyxl
import os
from .utils_translator import translate_batch
from app.services.terminology_service import (
    get_terminology_options,
    preprocess_texts,
    preprocess_texts_with_categories,
    postprocess_texts,
)
from app.database import SessionLocal
from .translator_xlsx_ooxml import translate_xlsx_ooxml

def is_translatable(cell_value):
    """判断单元格内容是否需要翻译：只翻译纯文本"""
    if cell_value is None or not isinstance(cell_value, str):
        return False
    cell_value = cell_value.strip()
    if not cell_value:
        return False
    if cell_value.startswith('=') or (cell_value.replace('.', '', 1).isdigit() and cell_value.count('.') < 2):
        return False
    return True

def translate_xlsx_direct(input_path, output_path, src_lang, tgt_lang, engine="deepseek", user_id: int | None = None, **kwargs):
    """Translate an XLSX file using openpyxl, preserving formatting and structure."""
    # 优先使用 OOXML 级处理，最大限度保留图形/形状/格式
    cat_ids = kwargs.get('category_ids')
    try:
        return translate_xlsx_ooxml(input_path, output_path, src_lang, tgt_lang, engine=engine, user_id=user_id, category_ids=cat_ids)
    except Exception:
        # 回退到 openpyxl 方案
        pass
    wb = openpyxl.load_workbook(input_path)
    total_token_count = 0
    total_character_count = 0
    
    all_texts_to_translate = []
    cell_map = {}

    # First pass: Collect all unique translatable texts from cells and comments
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if is_translatable(cell.value):
                    text = cell.value
                    total_character_count += len(text)
                    if text not in cell_map:
                        cell_map[text] = []
                        all_texts_to_translate.append(text)
                    cell_map[text].append(cell)
        
        # 收集单元格批注（兼容 openpyxl 的 worksheet 无 comments 属性场景）
        for row in ws.iter_rows():
            for cell in row:
                cmt = getattr(cell, 'comment', None)
                if cmt and getattr(cmt, 'text', None) and is_translatable(cmt.text):
                    text = cmt.text
                    total_character_count += len(text)
                    if text not in cell_map:
                        cell_map[text] = []
                        all_texts_to_translate.append(text)
                    cell_map[text].append(cmt)

    # Translate all collected unique texts in one batch
    translations = {}
    if all_texts_to_translate:
        db = SessionLocal()
        try:
            options = get_terminology_options(db)
            if options.get("terminology_enabled", True):
                cat_ids = kwargs.get("category_ids", None)
                if options.get("categories_enabled", True):
                    # 规则统一：启用分类但未选择分类(为空或未传)时，不应用术语
                    if not cat_ids or (isinstance(cat_ids, list) and len(cat_ids) == 0):
                        processed_texts, mappings = all_texts_to_translate, [{} for _ in all_texts_to_translate]
                    else:
                        processed_texts, mappings = preprocess_texts_with_categories(
                            db, all_texts_to_translate, src_lang, tgt_lang, cat_ids,
                            case_sensitive=bool(options.get("case_sensitive", False)), user_id=user_id
                        )
                else:
                    processed_texts, mappings = preprocess_texts(
                        db, all_texts_to_translate, src_lang, tgt_lang,
                        case_sensitive=bool(options.get("case_sensitive", False)), user_id=user_id
                    )
            else:
                processed_texts, mappings = all_texts_to_translate, [{} for _ in all_texts_to_translate]

            # 并行批量翻译，提升性能
            _res = translate_batch(processed_texts, src_lang, tgt_lang, engine=engine, **kwargs)
            if isinstance(_res, tuple) and len(_res) >= 2:
                translated_texts, _tokens = _res[0], _res[1]
                try:
                    total_token_count += int(_tokens or 0)
                except Exception:
                    pass
            else:
                translated_texts = _res
            if options.get("terminology_enabled", True):
                translated_texts = postprocess_texts(translated_texts, mappings)
            translations = dict(zip(all_texts_to_translate, translated_texts))
        finally:
            db.close()

    # Second pass: Write back translations
    for original_text, translated_text in translations.items():
        for item in cell_map[original_text]:
            if isinstance(item, openpyxl.cell.cell.Cell):
                item.value = translated_text
            elif isinstance(item, openpyxl.comments.Comment):
                item.text = translated_text

    # Save the translated workbook
    wb.save(output_path)
    
    # Return metadata
    return {
        "token_count": total_token_count,
        "character_count": total_character_count
    }

# Alias for compatibility
translate_xlsx = translate_xlsx_direct